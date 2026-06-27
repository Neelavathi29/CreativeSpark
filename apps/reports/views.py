from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Avg
import csv
from io import BytesIO, StringIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill
from apps.ideas.models import StartupIdea
from apps.evaluation.models import Evaluation


@login_required
def report_dashboard(request):
    total_ideas = StartupIdea.objects.count()
    status_counts = {}
    for s, label in StartupIdea.STATUS_CHOICES:
        count = StartupIdea.objects.filter(status=s).count()
        if count > 0:
            status_counts[label] = count

    industry_counts = {}
    for ind, label in StartupIdea.INDUSTRY_CHOICES:
        count = StartupIdea.objects.filter(industry=ind).count()
        if count > 0:
            industry_counts[label] = count

    avg_scores = Evaluation.objects.aggregate(
        avg_innovation=Avg("innovation_score"),
        avg_feasibility=Avg("feasibility_score"),
        avg_market=Avg("market_potential"),
        avg_scalability=Avg("scalability_score"),
        avg_overall=Avg("overall_rating"),
    )

    context = {
        "total_ideas": total_ideas,
        "status_counts": status_counts,
        "industry_counts": industry_counts,
        "avg_scores": avg_scores,
    }
    return render(request, "reports/report_dashboard.html", context)


@login_required
def export_pdf(request, pk=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    if pk:
        idea = get_object_or_404(StartupIdea, pk=pk)
        evaluation = Evaluation.objects.filter(idea=idea).first()

        elements.append(
            Paragraph("Startup Evaluation Report", styles["Title"])
        )
        elements.append(Spacer(1, 20))
        elements.append(
            Paragraph(f"Startup: {idea.startup_name}", styles["Heading2"])
        )
        elements.append(
            Paragraph(f"Founder: {idea.founder_name}", styles["Normal"])
        )
        elements.append(
            Paragraph(
                f"Industry: {idea.get_industry_display()}", styles["Normal"]
            )
        )
        elements.append(
            Paragraph(f"Status: {idea.get_status_display()}", styles["Normal"])
        )
        elements.append(Spacer(1, 10))

        if evaluation:
            elements.append(Paragraph("Evaluation Scores", styles["Heading2"]))
            score_data = [
                ["Metric", "Score"],
                ["Innovation", str(evaluation.innovation_score)],
                ["Feasibility", str(evaluation.feasibility_score)],
                ["Market Potential", str(evaluation.market_potential)],
                ["Scalability", str(evaluation.scalability_score)],
                ["Risk Assessment", str(evaluation.risk_score)],
                ["Overall Rating", str(evaluation.overall_rating)],
            ]
            table = Table(score_data)
            table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor("#6C63FF"),
                        ),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 12),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                        (
                            "BACKGROUND",
                            (0, 1),
                            (-1, -1),
                            colors.HexColor("#F8F9FA"),
                        ),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, 20))

            elements.append(
                Paragraph(
                    f"Recommendation: {evaluation.incubation_recommendation}",
                    styles["Heading3"],
                )
            )
            elements.append(Spacer(1, 10))
            elements.append(
                Paragraph(
                    "<b>Strengths:</b><br/>"
                    + evaluation.strengths.replace(chr(10), "<br/>"),
                    styles["Normal"],
                )
            )
            elements.append(Spacer(1, 10))
            elements.append(
                Paragraph(
                    "<b>Improvement Suggestions:</b><br/>"
                    + evaluation.improvement_suggestions.replace(chr(10), "<br/>"),
                    styles["Normal"],
                )
            )
    else:
        elements.append(
            Paragraph("Creative Spark - All Startups Report", styles["Title"])
        )
        elements.append(Spacer(1, 20))
        ideas = StartupIdea.objects.all()

        if ideas:
            data = [["Name", "Founder", "Industry", "Status", "Rating"]]
            for idea in ideas:
                ev = Evaluation.objects.filter(idea=idea).first()
                rating = str(ev.overall_rating) if ev else "N/A"
                data.append(
                    [
                        idea.startup_name,
                        idea.founder_name,
                        idea.get_industry_display(),
                        idea.get_status_display(),
                        rating,
                    ]
                )

            table = Table(data)
            table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor("#6C63FF"),
                        ),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    filename = f"startup_report_{pk if pk else 'all'}.pdf"
    response = HttpResponse(
        buffer, content_type="application/pdf"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_excel(request, pk=None):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Startup Report"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(
        start_color="6C63FF", end_color="6C63FF", fill_type="solid"
    )

    if pk:
        idea = get_object_or_404(StartupIdea, pk=pk)
        evaluation = Evaluation.objects.filter(idea=idea).first()

        ws.cell(row=1, column=1, value="Startup Evaluation Report").font = (
            Font(bold=True, size=14)
        )
        ws.merge_cells("A1:C1")

        headers = ["Metric", "Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        info = [
            ("Startup Name", idea.startup_name),
            ("Founder", idea.founder_name),
            ("Industry", idea.get_industry_display()),
            ("Status", idea.get_status_display()),
            (
                "Problem",
                (
                    idea.problem_statement[:100]
                    if idea.problem_statement
                    else "N/A"
                ),
            ),
        ]
        for i, (label, value) in enumerate(info, 4):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

        if evaluation:
            row = 10
            ws.cell(row=row, column=1, value="Evaluation Scores").font = Font(
                bold=True, size=12
            )
            ws.merge_cells(f"A{row}:B{row}")
            row += 1
            for col, header in enumerate(["Metric", "Score"], 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            scores = [
                ("Innovation", evaluation.innovation_score),
                ("Feasibility", evaluation.feasibility_score),
                ("Market Potential", evaluation.market_potential),
                ("Scalability", evaluation.scalability_score),
                ("Risk", evaluation.risk_score),
                ("Overall Rating", evaluation.overall_rating),
            ]
            for label, score in scores:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=score)
                row += 1
    else:
        ws.cell(row=1, column=1, value="All Startups Report").font = Font(
            bold=True, size=14
        )
        ws.merge_cells("A1:E1")

        headers = ["Name", "Founder", "Industry", "Status", "Rating"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        ideas = StartupIdea.objects.all()
        for i, idea in enumerate(ideas, 4):
            ev = Evaluation.objects.filter(idea=idea).first()
            rating = str(ev.overall_rating) if ev else "N/A"
            ws.cell(row=i, column=1, value=idea.startup_name)
            ws.cell(row=i, column=2, value=idea.founder_name)
            ws.cell(row=i, column=3, value=idea.get_industry_display())
            ws.cell(row=i, column=4, value=idea.get_status_display())
            ws.cell(row=i, column=5, value=rating)

    wb.save(output)
    output.seek(0)
    filename = f"startup_report_{'all' if not pk else pk}.xlsx"
    response = HttpResponse(
        output,
        content_type=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_csv(request, pk=None):
    output = StringIO()
    writer = csv.writer(output)

    if pk:
        idea = get_object_or_404(StartupIdea, pk=pk)
        evaluation = Evaluation.objects.filter(idea=idea).first()
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Startup Name", idea.startup_name])
        writer.writerow(["Founder", idea.founder_name])
        writer.writerow(["Industry", idea.get_industry_display()])
        writer.writerow(["Status", idea.get_status_display()])
        if evaluation:
            writer.writerow([])
            writer.writerow(["Evaluation Scores"])
            writer.writerow(["Innovation", evaluation.innovation_score])
            writer.writerow(["Feasibility", evaluation.feasibility_score])
            writer.writerow(["Market Potential", evaluation.market_potential])
            writer.writerow(["Scalability", evaluation.scalability_score])
            writer.writerow(["Risk", evaluation.risk_score])
            writer.writerow(["Overall Rating", evaluation.overall_rating])
    else:
        writer.writerow(["Name", "Founder", "Industry", "Status", "Rating"])
        ideas = StartupIdea.objects.all()
        for idea in ideas:
            ev = Evaluation.objects.filter(idea=idea).first()
            rating = str(ev.overall_rating) if ev else "N/A"
            writer.writerow([
                idea.startup_name,
                idea.founder_name,
                idea.get_industry_display(),
                idea.get_status_display(),
                rating,
            ])

    output.seek(0)
    filename = f"startup_report_{'all' if not pk else pk}.csv"
    response = HttpResponse(output, content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
